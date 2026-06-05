"""Build a corrected copy of the source xlsx that removes the duplicate rows
per the user's decision (2026-05-21).

Resolution:
  - 164 conflict empCodes -> BLUE_COLLAR: remove their row from
    'Production-White Collar' (the WC tab). Their row in the BC tab stays.
  - 2 conflict empCodes -> WHITE_COLLAR (keep in Production-White Collar,
    remove from BC tab):
      1800046  SYOJI RAM JAT          (remove from 'Production-Helper blue collor')
      1801105  Jagdish Prasad Meena   (remove from 'Production-Helper blue collor')

Source : C:\\Users\\Dinesh\\Downloads\\self_assessment_-Jaipur_EMployee_Details-Main_list1.xlsx
Output : C:\\Users\\Dinesh\\Downloads\\self_assessment_-Jaipur_EMployee_Details-Main_list1_FIXED.xlsx

Uses openpyxl so cell formatting / column widths / header styles are preserved.
"""
import shutil
from openpyxl import load_workbook

SRC = r"C:\Users\Dinesh\Downloads\self_assessment_-Jaipur_EMployee_Details-Main_list1.xlsx"
DST = r"C:\Users\Dinesh\Downloads\self_assessment_-Jaipur_EMployee_Details-Main_list1_FIXED.xlsx"

KEEP_WC = {"1800046", "1801105"}

# 166 conflicts: 164 → BC (remove from WC tab), 2 → WC (remove from BC tab)
GROUP1_PHB_x_PWC = [  # in 'Production-Helper blue collor' AND 'Production-White Collar'
    "1801105","1802153","1800056","1801872","1800169","1801167","1801368","1801776",
    "1802258","1802333","1801745","1802159","1802334","1802192","1800094","1802337",
    "1801416","1802006","1801363","1802232","1802222","1802299","1800133","1800106",
    "1802217","1801587","1802385","1800221","1800282","1800171","1800989","1802361",
    "1802284","1801180","1801919","1801841","1802253","1800144","1802075","1800168",
    "1802288","1802387","1802145","1801079","1800557","1801187","1800060","1800979",
    "1800134","1801748","1802188","1802328","1800135","1800160","1800110","1801840",
    "1802324","1802384","1801078","1801391","1801912","1802049","1801607","1800050",
    "1801642","1800120","1802371","1802289","1802264","1800153","1802147","1801735",
    "1801952","1800082","1802363","1802009","1802152","1802073","1801544","1802141",
    "1800137","1800112","1801747","1802240","1802233","1802099","1802090","1802377",
    "1801878","1800496","1800113","1802119","1801873","1800115","1801846","1802269",
    "1802375","1801053","1802118","1801780","1802360","1801892","1800978","1801534",
    "1800154","1801947","1802365","1801924","1800117","1801728","1801746","1801025",
    "1800046","1802383","1800143","1802254","1802287","1801427","1802345","1801938",
    "1801545","1801853","1800474","1802281","1802369",
]
GROUP2_DHB_x_PWC = [  # in 'Distribution-Helper blue collor' AND 'Production-White Collar'
    "1801219","1801220","1801221","1801781","1801658","1801225","1801226","1801227",
    "1802374","1802378","1801232","1801198","1802040","1801862","1801236","1801237",
    "1801238","1801961","1801240","1801241","1801242","1801243","1801767","1802331",
    "1801246","1801248","1802059","1802244","1801252","1801291","1802355","1801253",
    "1801254","1802339","1802057","1802025","1801656","1802069","1801768","1801259",
    "1801260",
]

# Build deletion plan: per tab name -> set of empCodes to delete
to_delete = {
    "Production-White Collar":          set(),  # everyone in group1 + group2 EXCEPT KEEP_WC
    "Production-Helper blue collor":    set(),  # only KEEP_WC people who appear here
    "Distribution-Helper blue collor":  set(),  # only KEEP_WC people who appear here
}
for code in GROUP1_PHB_x_PWC:
    if code in KEEP_WC:
        # remove from BC tab, keep in WC
        to_delete["Production-Helper blue collor"].add(code)
    else:
        # remove from WC tab, keep in BC
        to_delete["Production-White Collar"].add(code)
for code in GROUP2_DHB_x_PWC:
    if code in KEEP_WC:
        to_delete["Distribution-Helper blue collor"].add(code)
    else:
        to_delete["Production-White Collar"].add(code)

print("Planned row deletions per tab:")
for tab, codes in to_delete.items():
    print(f"  {tab!r}: {len(codes)} rows")

shutil.copyfile(SRC, DST)
wb = load_workbook(DST)

total_deleted = 0
for tab_name, codes_to_remove in to_delete.items():
    if not codes_to_remove:
        continue
    if tab_name not in wb.sheetnames:
        # try fuzzy match (the actual tabs sometimes have trailing spaces)
        match = None
        for s in wb.sheetnames:
            if s.strip() == tab_name.strip():
                match = s
                break
        if not match:
            print(f"!! Tab not found: {tab_name!r}, sheets are {wb.sheetnames}")
            continue
        ws = wb[match]
    else:
        ws = wb[tab_name]

    # Walk rows from bottom-up so deletions don't shift our index.
    # Identify header row first.
    header_row = 1
    first_cell = (ws.cell(row=1, column=1).value or "")
    if str(first_cell).strip().lower() != "empcode":
        header_row = 0  # no header

    rows_to_delete_idx = []
    for row_idx in range(header_row + 1, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=1).value
        emp = str(cell).strip() if cell is not None else ""
        if emp in codes_to_remove:
            rows_to_delete_idx.append(row_idx)

    # Delete bottom-up
    for r in sorted(rows_to_delete_idx, reverse=True):
        ws.delete_rows(r, 1)
    total_deleted += len(rows_to_delete_idx)
    print(f"  Tab '{ws.title}': deleted {len(rows_to_delete_idx)} rows")

wb.save(DST)
print(f"\nTotal rows deleted: {total_deleted}  (expected: 164+2 = 166)")
print(f"Output saved to: {DST}")
